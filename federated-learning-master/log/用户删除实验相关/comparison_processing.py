import re
import matplotlib.pyplot as plt
import matplotlib
import pandas as pd
import numpy as np
from collections import defaultdict
from openpyxl import Workbook
from matplotlib.ticker import MaxNLocator
from datetime import datetime

# 设置中文字体和样式
plt.style.use('seaborn')
font = {'family': 'SimHei', 'size': 12, 'weight': 'bold'}
matplotlib.rc('font', **font)
matplotlib.rcParams['axes.titlepad'] = 20


class LogParser:
    def __init__(self):
        self.all_stages = []
        self.current_stage = []
        self.active_clients = set(range(10))
        self.global_epoch_counter = 0
        self.stage_counter = 0
        self.test_accuracies = []
        self.current_epoch = -1
        self.start_time = None

    def parse_datetime(self, line):
        """解析日志时间戳"""
        time_str = line.split(' - ')[0].strip()
        return datetime.strptime(time_str, '%Y-%m-%d %H:%M:%S,%f')

    def parse_log(self, log_content):
        """主解析函数"""
        lines = log_content.split('\n')
        for i, line in enumerate(lines):
            if not line.strip():
                continue

            # 初始化开始时间
            if self.start_time is None and re.match(r'^\d{4}-\d{2}-\d{2}', line):
                self.start_time = self.parse_datetime(line)

            # 解析活跃客户端状态
            if "Active clients participation:" in line:
                self._parse_active_clients(line)

            # 处理用户停用事件
            if "Deactivated user" in line:
                self._handle_deactivation(line)

            # 解析epoch信息
            if "epoch:" in line:
                self._parse_epoch(line)

            # 解析参与客户端
            if "Clients participation:" in line:
                self._parse_participants(line)

            # 解析评分数据
            if "Clients Score" in line:
                self._parse_scores(line, lines[i + 1] if i + 1 < len(lines) else "")

        # 填充准确率数据
        self._fill_accuracies()
        return self.all_stages

    def _parse_active_clients(self, line):
        """解析活跃客户端列表"""
        clients_str = line.split("{")[1].split("}")[0]
        self.active_clients = set()
        for pair in clients_str.split(", "):
            try:
                client_id = int(pair.split(":")[0])
                self.active_clients.add(client_id)
            except (ValueError, IndexError):
                continue

    def _handle_deactivation(self, line):
        """处理客户端停用事件"""
        deactivated_match = re.search(r'Deactivated user (-?\d+)', line)
        accuracy_match = re.search(r'Test Accuracy: ([\d.]+)%', line)

        deactivated = int(deactivated_match.group(1)) if deactivated_match else None
        accuracy = float(accuracy_match.group(1)) if accuracy_match else None

        if deactivated is not None and deactivated >= 0:
            self.active_clients.discard(deactivated)

        if accuracy is not None:
            self.test_accuracies.append(accuracy)

        # 提交当前阶段
        if self.current_stage:
            self.all_stages.append({
                "stage_id": self.stage_counter,
                "epochs": self.current_stage.copy(),
                "deactivated": deactivated,
                "global_start_epoch": self.global_epoch_counter - len(self.current_stage),
                "test_accuracy": accuracy
            })
            self.stage_counter += 1
            self.current_stage = []
            self.current_epoch = -1

    def _parse_epoch(self, line):
        """解析epoch信息"""
        epoch_match = re.search(r'epoch: (\d+)', line)
        if epoch_match:
            self.current_epoch = int(epoch_match.group(1))
            if len(self.current_stage) <= self.current_epoch:
                self.current_stage.append({
                    "local_epoch": self.current_epoch,
                    "global_epoch": self.global_epoch_counter,
                    "participants": [],
                    "scores_acc": defaultdict(float),
                    "scores_g_like": defaultdict(float),
                    "active_clients": self.active_clients.copy(),
                    "test_accuracy": None,
                    "timestamp": self.parse_datetime(line) if re.match(r'^\d{4}-\d{2}-\d{2}', line) else None
                })
                self.global_epoch_counter += 1

    def _parse_participants(self, line):
        """解析参与客户端"""
        if self.current_epoch == -1:
            return

        try:
            participants = list(map(int, re.findall(r'-?\d+', line.split(": [")[1].rstrip(']'))))
            self.current_stage[self.current_epoch]["participants"] = [c for c in participants if
                                                                      c in self.active_clients]
        except (IndexError, ValueError):
            pass

    def _parse_scores(self, line, next_line):
        """解析评分数据"""
        if self.current_epoch == -1:
            return

        score_type = None
        if "Clients Score using acc:" in line:
            score_type = "acc"
        elif "Clients Score using glike:" in line:
            score_type = "g_like"
        elif "Clients Score:" in line and "using method:" in next_line:
            method_match = re.search(r'using method: (\w+)', next_line)
            if method_match:
                score_type = method_match.group(1).lower()

        if score_type:
            try:
                scores = list(map(float, re.findall(r'-?\d+\.\d+', line.split(": [")[1].rstrip(']'))))
                for client_id, score in enumerate(scores):
                    if client_id in self.active_clients:
                        self.current_stage[self.current_epoch][f"scores_{score_type}"][client_id] = score
            except (IndexError, ValueError):
                pass

    def _fill_accuracies(self):
        """填充准确率数据"""
        accuracy_index = 0
        for stage in self.all_stages:
            if stage["epochs"] and accuracy_index < len(self.test_accuracies):
                stage["epochs"][-1]["test_accuracy"] = self.test_accuracies[accuracy_index]
                accuracy_index += 1


class ResultExporter:
    @staticmethod
    def export_to_excel(all_stages, filename="fl_results.xlsx"):
        wb = Workbook()

        # 阶段汇总表
        summary_ws = wb.active
        summary_ws.title = "阶段汇总"
        summary_headers = ["阶段ID", "停用客户端", "Epoch数", "起始全局Epoch",
                           "终止全局Epoch", "初始准确率", "最终准确率", "准确率变化"]
        summary_ws.append(summary_headers)

        # 详细数据表
        detail_ws = wb.create_sheet("详细数据")
        detail_headers = ["阶段ID", "本地Epoch", "全局Epoch", "参与客户端",
                          "活跃客户端数", "测试准确率", "Top 3(acc)", "Top 3(glike)", "时间戳"]
        detail_ws.append(detail_headers)

        # 客户端分析表
        client_ws = wb.create_sheet("客户端分析")
        client_headers = ["客户端ID", "总参与次数", "ACC总分数", "GLIKE总分数",
                          "ACC平均分", "GLIKE平均分", "最后活跃阶段"]
        client_ws.append(client_headers)

        # 客户端统计
        client_stats = {client: {
            "acc_scores": [],
            "glike_scores": [],
            "participations": 0,
            "last_active": -1
        } for client in range(10)}

        # 填充数据
        for stage in all_stages:
            # 阶段汇总
            start_acc = stage["epochs"][0]["test_accuracy"] if stage["epochs"] else None
            end_acc = stage["epochs"][-1]["test_accuracy"] if stage["epochs"] else None
            summary_ws.append([
                stage["stage_id"],
                stage["deactivated"],
                len(stage["epochs"]),
                stage["global_start_epoch"],
                stage["global_start_epoch"] + len(stage["epochs"]) - 1,
                start_acc,
                end_acc,
                end_acc - start_acc if (start_acc and end_acc) else None
            ])

            # 详细数据
            for epoch in stage["epochs"]:
                # 更新客户端统计
                for client in epoch["participants"]:
                    client_stats[client]["participations"] += 1
                    if client in epoch["scores_acc"]:
                        client_stats[client]["acc_scores"].append(epoch["scores_acc"][client])
                    if client in epoch["scores_g_like"]:
                        client_stats[client]["glike_scores"].append(epoch["scores_g_like"][client])
                    client_stats[client]["last_active"] = max(
                        client_stats[client]["last_active"],
                        stage["stage_id"]
                    )

                # 获取Top 3客户端
                top_acc = sorted(
                    [(k, v) for k, v in epoch["scores_acc"].items()],
                    key=lambda x: x[1],
                    reverse=True
                )[:3]
                top_glike = sorted(
                    [(k, v) for k, v in epoch["scores_g_like"].items()],
                    key=lambda x: x[1],
                    reverse=True
                )[:3]

                detail_ws.append([
                    stage["stage_id"],
                    epoch["local_epoch"],
                    epoch["global_epoch"],
                    ",".join(map(str, epoch["participants"])),
                    len(epoch["active_clients"]),
                    epoch["test_accuracy"],
                    "; ".join([f"{k}:{v:.2f}" for k, v in top_acc]),
                    "; ".join([f"{k}:{v:.2f}" for k, v in top_glike]),
                    epoch["timestamp"].strftime('%Y-%m-%d %H:%M:%S') if epoch["timestamp"] else ""
                ])

        # 客户端分析数据
        for client in range(10):
            acc_scores = client_stats[client]["acc_scores"]
            glike_scores = client_stats[client]["glike_scores"]

            client_ws.append([
                client,
                client_stats[client]["participations"],
                sum(acc_scores) if acc_scores else 0,
                sum(glike_scores) if glike_scores else 0,
                np.mean(acc_scores) if acc_scores else None,
                np.mean(glike_scores) if glike_scores else None,
                client_stats[client]["last_active"]
            ])

        # 设置列宽
        for sheet in wb:
            for column in sheet.columns:
                max_length = max(len(str(cell.value)) for cell in column) + 2
                sheet.column_dimensions[column[0].column_letter].width = min(max_length, 30)

        wb.save(filename)
        print(f"数据已导出到 {filename}")


class Visualizer:
    @staticmethod
    def visualize(all_stages):
        # 合并所有epoch数据
        all_epochs = []
        for stage in all_stages:
            all_epochs.extend(stage["epochs"])

        # 1. 客户端参与热力图
        plt.figure(figsize=(18, 8))
        participation_matrix = np.zeros((10, len(all_epochs)))

        for epoch_idx, epoch in enumerate(all_epochs):
            for client in epoch["participants"]:
                if 0 <= client < 10:  # 确保客户端ID有效
                    participation_matrix[client, epoch_idx] += 1

        plt.imshow(participation_matrix, cmap="YlGnBu", aspect="auto", interpolation="none")
        plt.colorbar(label="参与次数", pad=0.02)

        # 标记阶段分隔线
        for stage in all_stages[1:]:
            plt.axvline(x=stage["global_start_epoch"] - 0.5, color='r', linestyle='--', linewidth=1)

        plt.xlabel("全局Epoch", fontsize=14, fontweight='bold')
        plt.ylabel("客户端ID", fontsize=14, fontweight='bold')
        plt.title("客户端参与热力图（红色虚线表示阶段分隔）", fontsize=16, pad=20)
        plt.yticks(range(10))
        plt.grid(False)
        plt.tight_layout()
        plt.show()

        # 2. 双评分对比图
        Visualizer._plot_score_comparison(all_epochs)

        # 3. 活跃客户端与准确率趋势
        Visualizer._plot_active_clients_accuracy(all_stages, all_epochs)

        # 4. 客户端贡献分析
        Visualizer._plot_client_contributions(all_epochs)

        # 5. 训练时间分析
        Visualizer._plot_training_timeline(all_epochs)

    @staticmethod
    def _plot_score_comparison(all_epochs):
        """绘制两种评分对比图"""
        plt.figure(figsize=(12, 6))

        # 计算每个epoch的平均分
        acc_means = []
        glike_means = []
        for epoch in all_epochs:
            acc_scores = list(epoch["scores_acc"].values())
            glike_scores = list(epoch["scores_g_like"].values())
            acc_means.append(np.mean(acc_scores) if acc_scores else 0)
            glike_means.append(np.mean(glike_scores) if glike_scores else 0)

        plt.plot(range(len(all_epochs)), acc_means, label='ACC平均分', marker='o')
        plt.plot(range(len(all_epochs)), glike_means, label='GLIKE平均分', marker='s')
        plt.xlabel("全局Epoch")
        plt.ylabel("平均分数")
        plt.title("ACC与GLIKE评分对比")
        plt.legend()
        plt.grid(True)
        plt.show()

    @staticmethod
    def _plot_active_clients_accuracy(all_stages, all_epochs):
        """绘制活跃客户端与准确率趋势图"""
        fig, ax1 = plt.subplots(figsize=(15, 6))

        # 活跃客户端数
        active_counts = [len(e["active_clients"]) for e in all_epochs]
        ax1.plot(range(len(all_epochs)), active_counts, 'b-o', label='活跃客户端数')
        ax1.set_xlabel("全局Epoch")
        ax1.set_ylabel("活跃客户端数", color='b')
        ax1.tick_params(axis='y', labelcolor='b')
        ax1.set_ylim(0, 10)

        # 测试准确率
        ax2 = ax1.twinx()
        accuracies = [e["test_accuracy"] for e in all_epochs if e["test_accuracy"] is not None]
        ax2.plot(range(len(accuracies)), accuracies, 'r--s', label='测试准确率')
        ax2.set_ylabel("准确率 (%)", color='r')
        ax2.tick_params(axis='y', labelcolor='r')
        ax2.set_ylim(70, 100)

        # 标记阶段分隔
        for stage in all_stages[1:]:
            ax1.axvline(x=stage["global_start_epoch"], color='g', linestyle=':', alpha=0.5)
            ax1.text(stage["global_start_epoch"], 9,
                     f"停用 {stage['deactivated']}", rotation=90, fontsize=10)

        plt.title("活跃客户端数与模型准确率趋势")
        fig.legend(loc="upper right")
        plt.grid(True)
        plt.show()

    @staticmethod
    def _plot_client_contributions(all_epochs):
        """绘制客户端贡献分析"""
        plt.figure(figsize=(14, 6))

        # 计算累计贡献
        acc_totals = defaultdict(float)
        glike_totals = defaultdict(float)
        participations = defaultdict(int)

        for epoch in all_epochs:
            for client in epoch["participants"]:
                participations[client] += 1
                if client in epoch["scores_acc"]:
                    acc_totals[client] += epoch["scores_acc"][client]
                if client in epoch["scores_g_like"]:
                    glike_totals[client] += epoch["scores_g_like"][client]

        clients = sorted(participations.keys())
        x = np.arange(len(clients))
        width = 0.35

        # 绘制双柱状图
        acc_bars = plt.bar(x - width / 2, [acc_totals[c] for c in clients],
                           width, label='ACC总分', color='skyblue')
        glike_bars = plt.bar(x + width / 2, [glike_totals[c] for c in clients],
                             width, label='GLIKE总分', color='orange')

        # 添加数据标签
        for bars in [acc_bars, glike_bars]:
            for bar in bars:
                height = bar.get_height()
                plt.text(bar.get_x() + bar.get_width() / 2., height,
                         f'{height:.1f}',
                         ha='center', va='bottom', fontsize=8)

        plt.xlabel("客户端ID")
        plt.ylabel("累计分数")
        plt.title("客户端累计贡献对比")
        plt.xticks(x, clients)
        plt.legend()
        plt.grid(True, axis='y')
        plt.tight_layout()
        plt.show()

    @staticmethod
    def _plot_training_timeline(all_epochs):
        """绘制训练时间线"""
        if not all_epochs or not all_epochs[0].get("timestamp"):
            return

        timestamps = [e["timestamp"] for e in all_epochs if e.get("timestamp")]
        durations = [(ts - timestamps[0]).total_seconds() / 60 for ts in timestamps]

        plt.figure(figsize=(12, 4))
        plt.plot(durations, range(len(durations)), 'b-')
        plt.xlabel("训练时间 (分钟)")
        plt.ylabel("全局Epoch")
        plt.title("训练进度时间线")
        plt.grid(True)
        plt.tight_layout()
        plt.show()


if __name__ == "__main__":
    # 使用示例
    with open("log_delete_experiment_comparison_20250406_001555.txt", "r", encoding='utf-8') as f:
        log_content = f.read()

    parser = LogParser()
    all_stages = parser.parse_log(log_content)

    ResultExporter.export_to_excel(all_stages, "federated_learning_results.xlsx")
    Visualizer.visualize(all_stages)