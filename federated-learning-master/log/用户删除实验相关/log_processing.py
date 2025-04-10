import re
import matplotlib.pyplot as plt
import matplotlib
import pandas as pd
import numpy as np
from collections import defaultdict
from openpyxl import Workbook
from matplotlib.ticker import MaxNLocator

# 设置中文字体和样式
plt.style.use('seaborn')
font = {'family': 'SimHei', 'size': 12, 'weight': 'bold'}
matplotlib.rc('font', **font)
matplotlib.rcParams['axes.titlepad'] = 20


def parse_log(log_content):
    all_stages = []
    current_stage = []
    active_clients = set(range(10))
    current_epoch = -1
    test_accuracies = []
    global_epoch_counter = 0  # 全局epoch计数器
    stage_counter = 0

    for line in log_content.split('\n'):
        # 检测新的训练阶段
        if "Deactivated user" in line:
            if current_stage:
                all_stages.append({
                    "stage_id": stage_counter,
                    "epochs": current_stage.copy(),
                    "deactivated": int(re.search(r'Deactivated user (\d+)', line).group(1)),
                    "global_start_epoch": global_epoch_counter - len(current_stage)
                })
                stage_counter += 1
                current_stage = []

            deactivated = int(re.search(r'Deactivated user (\d+)', line).group(1))
            active_clients.discard(deactivated)
            current_epoch = -1

        # 解析epoch
        if "epoch:" in line:
            current_epoch = int(re.search(r'epoch: (\d+)', line).group(1))
            if len(current_stage) <= current_epoch:
                current_stage.append({
                    "local_epoch": current_epoch,
                    "global_epoch": global_epoch_counter,
                    "participants": [],
                    "scores": defaultdict(float),
                    "active_clients": active_clients.copy(),
                    "test_accuracy": None
                })
                global_epoch_counter += 1

        # 解析参与客户端
        if "Clients participation:" in line:
            participants = list(map(int, re.findall(r'\d+', line.split(": [")[1].rstrip(']'))))
            current_stage[current_epoch]["participants"] = participants

        # 解析分数
        if "Clients Score:" in line:
            scores = list(map(float, re.findall(r'-?\d+\.\d+', line.split(": [")[1].rstrip(']'))))
            for client_id, score in enumerate(scores):
                if client_id in active_clients:
                    current_stage[current_epoch]["scores"][client_id] = score

        # 解析测试准确率
        if "Test Accuracy:" in line:
            accuracy = float(re.search(r'Test Accuracy: ([\d.]+)%', line).group(1))
            test_accuracies.append(accuracy)

    # 处理最后一个阶段
    if current_stage:
        all_stages.append({
            "stage_id": stage_counter,
            "epochs": current_stage,
            "deactivated": None,
            "global_start_epoch": global_epoch_counter - len(current_stage)
        })

    # 填充测试准确率
    accuracy_index = 0
    for stage in all_stages:
        for epoch in stage["epochs"]:
            if accuracy_index < len(test_accuracies):
                epoch["test_accuracy"] = test_accuracies[accuracy_index]
                accuracy_index += 1

    return all_stages


def export_to_excel(all_stages, filename="fl_results.xlsx"):
    wb = Workbook()

    # 阶段汇总表
    summary_ws = wb.active
    summary_ws.title = "阶段汇总"
    summary_headers = ["阶段ID", "停用客户端", "Epoch数", "起始全局Epoch",
                       "终止全局Epoch", "初始准确率", "最终准确率", "准确率变化"]
    summary_ws.append(summary_headers)

    # 详细数据表
    detail_ws = wb.create_sheet("详细数据")
    detail_headers = ["阶段ID", "本地Epoch", "全局Epoch", "参与客户端",
                      "活跃客户端数", "测试准确率", "Top 3客户端(分数)"]
    detail_ws.append(detail_headers)

    # 客户端分数表
    client_ws = wb.create_sheet("客户端分析")
    client_headers = ["客户端ID"] + [f"阶段{i}" for i in range(len(all_stages))] + ["总参与次数", "平均分数"]
    client_ws.append(client_headers)

    # 客户端统计字典
    client_stats = {client: {"scores": [], "participations": 0} for client in range(10)}

    for stage in all_stages:
        # 汇总表数据
        start_acc = stage["epochs"][0]["test_accuracy"] if stage["epochs"] else None
        end_acc = stage["epochs"][-1]["test_accuracy"] if stage["epochs"] else None
        summary_ws.append([
            stage["stage_id"],
            stage["deactivated"],
            len(stage["epochs"]),
            stage["global_start_epoch"],
            stage["global_start_epoch"] + len(stage["epochs"]) - 1,
            start_acc,
            end_acc,
            end_acc - start_acc if (start_acc and end_acc) else None
        ])

        # 详细数据
        for epoch in stage["epochs"]:
            # 更新客户端统计
            for client in epoch["participants"]:
                client_stats[client]["participations"] += 1
                client_stats[client]["scores"].append(epoch["scores"][client])

            # 获取分数Top 3客户端
            top_clients = sorted(epoch["scores"].items(), key=lambda x: x[1], reverse=True)[:3]
            top_str = "; ".join([f"{k}:{v:.2f}" for k, v in top_clients])

            detail_ws.append([
                stage["stage_id"],
                epoch["local_epoch"],
                epoch["global_epoch"],
                ",".join(map(str, epoch["participants"])),
                len(epoch["active_clients"]),
                epoch["test_accuracy"],
                top_str
            ])

    # 客户端分析数据
    for client in range(10):
        stage_scores = []
        for stage in all_stages:
            avg_score = np.mean([epoch["scores"][client] for epoch in stage["epochs"] if client in epoch["scores"]])
            stage_scores.append(avg_score if not np.isnan(avg_score) else None)

        client_ws.append([
            client,
            *[f"{x:.2f}" if x is not None else "" for x in stage_scores],
            client_stats[client]["participations"],
            np.mean(client_stats[client]["scores"]) if client_stats[client]["scores"] else None
        ])

    # 设置列宽
    for sheet in wb:
        for column in sheet.columns:
            max_length = max(len(str(cell.value)) for cell in column) + 2
            sheet.column_dimensions[column[0].column_letter].width = min(max_length, 30)

    wb.save(filename)
    print(f"数据已导出到 {filename}")


def visualize(all_stages):
    # 合并所有epoch数据
    all_epochs = []
    for stage in all_stages:
        all_epochs.extend(stage["epochs"])

    # 准备全局数据
    global_epochs = [e["global_epoch"] for e in all_epochs]
    active_counts = [len(e["active_clients"]) for e in all_epochs]
    accuracies = [e["test_accuracy"] for e in all_epochs]

    # 1. 客户端参与热力图
    plt.figure(figsize=(18, 8))
    participation_matrix = np.zeros((10, len(all_epochs)))

    for epoch_idx, epoch in enumerate(all_epochs):
        for client in epoch["participants"]:
            participation_matrix[client, epoch_idx] += 1

    plt.imshow(participation_matrix, cmap="YlGnBu", aspect="auto", interpolation="none")
    plt.colorbar(label="参与次数", pad=0.02)

    # 标记阶段分隔线
    for stage in all_stages[1:]:
        plt.axvline(x=stage["global_start_epoch"] - 0.5, color='r', linestyle='--', linewidth=1)

    plt.xlabel("全局Epoch", fontsize=14, fontweight='bold')
    plt.ylabel("客户端ID", fontsize=14, fontweight='bold')
    plt.title("客户端参与热力图（红色虚线表示阶段分隔）", fontsize=16, pad=20)
    plt.yticks(range(10))
    plt.grid(False)
    plt.tight_layout()
    plt.show()

    # 2. 分数演变雷达图（按阶段）
    plt.figure(figsize=(10, 10))
    ax = plt.subplot(111, polar=True)

    # 准备阶段数据
    stages = len(all_stages)
    angles = np.linspace(0, 2 * np.pi, stages, endpoint=False).tolist()

    # 选择前5个客户端展示
    for client in range(5):
        client_scores = []
        for stage in all_stages:
            scores = [epoch["scores"][client] for epoch in stage["epochs"] if client in epoch["scores"]]
            client_scores.append(np.mean(scores) if scores else 0)

        # 闭合雷达图
        values = client_scores + [client_scores[0]]
        angles_plot = angles + [angles[0]]

        ax.plot(angles_plot, values, linewidth=2, linestyle='solid',
                label=f"Client {client}", marker='o')
        ax.fill(angles_plot, values, alpha=0.1)

    ax.set_thetagrids(np.degrees(angles), labels=[f"阶段{i}" for i in range(stages)])
    plt.title("客户端贡献分数阶段变化（雷达图）", fontsize=16, pad=30)
    plt.legend(bbox_to_anchor=(1.2, 1))
    plt.tight_layout()
    plt.show()

    # 3. 活跃客户端与准确率双轴图
    fig, ax1 = plt.subplots(figsize=(15, 6))

    # 绘制活跃客户端数
    ax1.plot(global_epochs, active_counts, 'b-o', linewidth=2, markersize=8, label='活跃客户端数')
    ax1.set_xlabel("全局Epoch", fontsize=14)
    ax1.set_ylabel("活跃客户端数", color='b', fontsize=14)
    ax1.tick_params(axis='y', labelcolor='b')
    ax1.grid(True, linestyle='--', alpha=0.6)

    # 绘制准确率
    ax2 = ax1.twinx()
    ax2.plot(global_epochs, accuracies, 'r--s', linewidth=2, markersize=8, label='测试准确率')
    ax2.set_ylabel("准确率 (%)", color='r', fontsize=14)
    ax2.tick_params(axis='y', labelcolor='r')
    ax2.set_ylim(70, 100)

    # 添加阶段标记
    for stage in all_stages:
        if stage["deactivated"] is not None:
            ax1.axvline(x=stage["global_start_epoch"], color='g', linestyle=':',
                        linewidth=1, alpha=0.5)
            ax1.text(stage["global_start_epoch"], max(active_counts) * 0.9,
                     f"停用 {stage['deactivated']}", rotation=90, fontsize=10)

    plt.title("活跃客户端数与模型准确率变化", fontsize=16, pad=20)
    fig.legend(loc="upper right", bbox_to_anchor=(0.9, 0.9))
    plt.tight_layout()
    plt.show()

    # 4. 客户端累计贡献条形图
    plt.figure(figsize=(12, 6))

    # 计算总贡献
    total_scores = {client: 0 for client in range(10)}
    for epoch in all_epochs:
        for client, score in epoch["scores"].items():
            total_scores[client] += score

    clients = list(total_scores.keys())
    scores = list(total_scores.values())

    # 绘制条形图
    bars = plt.bar(clients, scores, color=plt.cm.tab20(np.arange(10) / 10))

    # 添加数据标签
    for bar in bars:
        height = bar.get_height()
        plt.text(bar.get_x() + bar.get_width() / 2., height,
                 f'{height:.1f}',
                 ha='center', va='bottom')

    plt.xlabel("客户端ID", fontsize=14)
    plt.ylabel("累计贡献分数", fontsize=14)
    plt.title("客户端累计贡献分数排名", fontsize=16)
    plt.xticks(clients)
    plt.grid(True, axis='y', linestyle='--', alpha=0.6)
    plt.tight_layout()
    plt.show()


# 使用示例
if __name__ == "__main__":
    # 读取日志文件
    with open("log_GTG-Shapley_low-quality remove_mnist_False_mlp_20250403_114354.txt", "r", encoding='utf-8') as f:
        log_content = f.read()

    # 解析数据
    all_stages = parse_log(log_content)

    # 导出Excel
    export_to_excel(all_stages)

    # 可视化
    visualize(all_stages)